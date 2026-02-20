# exporter_v2.py (LAZY IMPORTS)
# - Evita travar/atrasar a abertura do sistema: pandas/openpyxl só são importados quando o usuário manda exportar.
# - Mantém compatibilidade com o interface_inicial_v2.py (função export_bsdepara).
# - Join do físico aceita ID ou row_id (para cobrir casos antigos do manual).

from __future__ import annotations

import re
import sqlite3
from typing import List, Tuple, TYPE_CHECKING
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile
from db_utils_v2 import connect

if TYPE_CHECKING:  # só para type-checkers
    import pandas as pd  # noqa: F401


# =========================
# Helpers: Template headers
# =========================

def _find_sheet(wb, name: str) -> str:
    wanted = name.strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == wanted:
            return s
    return name


def _read_headers(wb, sheet_name: str) -> List[str]:
    ws = wb[_find_sheet(wb, sheet_name)]
    headers: List[str] = []
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=j).value
        headers.append("" if v is None else str(v).strip())
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def _ensure_sheet(wb, name: str, headers: List[str]):
    key = None
    wanted = name.strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == wanted:
            key = s
            break
    ws = wb[key] if key else wb.create_sheet(title=name)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    return ws


def _clear_data(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _write_df(ws, df, headers: List[str]):
    # garante ordem e colunas
    for h in headers:
        if h not in df.columns:
            df[h] = ""
    df = df[headers]

    _clear_data(ws)
    # append por linha é significativamente mais rápido do que escrever célula-a-célula.
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def _save_workbook_fast(
    wb,
    out_xlsx: str,
    *,
    compresslevel: int = 1,
    ultra_fast: bool = False,
) -> None:
    """
    Salva workbook com compressão DEFLATE mais leve para reduzir tempo de exportação.
    Mantém compatibilidade do arquivo final (.xlsx).
    """
    from openpyxl.writer.excel import ExcelWriter

    compression = ZIP_STORED if ultra_fast else ZIP_DEFLATED
    with ZipFile(
        out_xlsx,
        mode="w",
        compression=compression,
        allowZip64=True,
        compresslevel=(None if ultra_fast else int(compresslevel)),
    ) as archive:
        writer = ExcelWriter(wb, archive)
        writer.save()


# =========================
# Sobras (pendentes)
# =========================

FISICO_HEADERS = [
    "ID","FILIAL","DESC.FILIAL","CCUSTO","DESCR.CCUSTO","LOCAL","DESCR. LOCAL","NRBRM","INC.","DESCRICAO",
    "MARCA","MODELO","SERIE","DIMENSAO","CAPACIDADE","TAG","BEM ANTERIOR","CONDIC","QTD","FRAG"
]

CONTABIL_HEADERS = [
    "ID","COD. CONTA","DESCR. CONTA","FILIAL","DESC.FILIAL","CCUSTO","DESCR.CCUSTO","LOCAL","DESCR. LOCAL","NRBRM","INC.","DESCRICAO",
    "MARCA","MODELO","SERIE","DIMENSAO","CAPACIDADE","TAG","BEM ANTERIOR","QTD",
    "DT. AQUISIÇÃO","VLR. AQUISIÇÃO","DEP. ACUMULADA","VLR. RESIDUAL","FRAG","ST_CONCILIACAO"
]

FISICO_MAP = {
    "ID":"ID",
    "FILIAL":"FILIAL",
    "DESC.FILIAL":"DESC_FILIAL",
    "CCUSTO":"CCUSTO",
    "DESCR.CCUSTO":"DESCR_CCUSTO",
    "LOCAL":"LOCAL",
    "DESCR. LOCAL":"DESCR_LOCAL",
    "NRBRM":"NRBRM",
    "INC.":"INC",
    "DESCRICAO":"DESCRICAO",
    "MARCA":"MARCA",
    "MODELO":"MODELO",
    "SERIE":"SERIE",
    "DIMENSAO":"DIMENSAO",
    "CAPACIDADE":"CAPACIDADE",
    "TAG":"TAG",
    "BEM ANTERIOR":"BEM_ANTERIOR",
    "CONDIC":"CONDIC",
    "QTD":"QTD",
    "FRAG":"FRAG",
}

CONTABIL_MAP = {
    "ID":"ID",
    "COD. CONTA":"COD_CONTA",
    "DESCR. CONTA":"DESC_CONTA",
    "FILIAL":"FILIAL",
    "DESC.FILIAL":"DESC_FILIAL",
    "CCUSTO":"CCUSTO",
    "DESCR.CCUSTO":"DESCR_CCUSTO",
    "LOCAL":"LOCAL",
    "DESCR. LOCAL":"DESCR_LOCAL",
    "NRBRM":"NRBRM",
    "INC.":"INC",
    "DESCRICAO":"DESCRICAO",
    "MARCA":"MARCA",
    "MODELO":"MODELO",
    "SERIE":"SERIE",
    "DIMENSAO":"DIMENSAO",
    "CAPACIDADE":"CAPACIDADE",
    "TAG":"TAG",
    "BEM ANTERIOR":"BEM_ANTERIOR",
    "QTD":"QTD",
    "DT. AQUISIÇÃO":"DT_AQUISICAO",
    "VLR. AQUISIÇÃO":"VLR_AQUISICAO",
    "DEP. ACUMULADA":"DEP_ACUMULADA",
    "VLR. RESIDUAL":"VLR_RESIDUAL",
    "FRAG":"FRAG",
    "ST_CONCILIACAO":"ST_CONCILIACAO",
}


def _build_pending_df(con: sqlite3.Connection, base: str):
    import pandas as pd
    def _table_cols(table: str) -> set[str]:
        rows = con.execute(f"PRAGMA table_info({table});").fetchall()
        return {str(r[1]).strip().upper() for r in rows}

    def _select_expr(table: str, db_col: str, alias: str, table_cols_upper: set[str]) -> str:
        col = (db_col or "").strip()
        if col.upper() in table_cols_upper:
            return f'{table}."{col}" AS "{alias}"'
        return f'NULL AS "{alias}"'

    if base == "FIS":
        fis_cols = _table_cols("fisico")
        fis_select = ", ".join(
            _select_expr("f", FISICO_MAP[h], h, fis_cols)
            for h in FISICO_HEADERS
        )
        q = f"""
        SELECT {fis_select}
        FROM fisico f
        WHERE f.ID IS NOT NULL
          AND NOT EXISTS (
            SELECT 1 FROM conciliados c
            WHERE c.BASE='FIS' AND c.ID=f.ID
          )
        """
        return pd.read_sql_query(q, con)

    if base == "CTB":
        ctb_cols = _table_cols("contabil")
        ctb_select = ", ".join(
            _select_expr("t", CONTABIL_MAP[h], h, ctb_cols)
            for h in CONTABIL_HEADERS
            if h != "ST_CONCILIACAO"
        )
        q = f"""
        SELECT {ctb_select}, '' AS "ST_CONCILIACAO"
        FROM contabil t
        WHERE t.ID IS NOT NULL
          AND NOT EXISTS (
            SELECT 1 FROM conciliados c
            WHERE c.BASE='CTB' AND c.ID=t.ID
          )
        """
        out = pd.read_sql_query(q, con)

        # data dd/mm/aaaa
        if "DT. AQUISIÇÃO" in out.columns:
            dt = pd.to_datetime(out["DT. AQUISIÇÃO"], errors="coerce")
            out["DT. AQUISIÇÃO"] = dt.dt.strftime("%d/%m/%Y").fillna("")

        return out

    raise ValueError("base deve ser 'FIS' ou 'CTB'")


def _table_has_column(con: sqlite3.Connection, table: str, column: str) -> bool:
    try:
        rows = con.execute(f"PRAGMA table_info({table});").fetchall()
    except Exception:
        return False
    col_upper = (column or "").strip().upper()
    return any(str(r[1]).strip().upper() == col_upper for r in rows)


def _needs_fis_rowid_fallback(con: sqlite3.Connection) -> bool:
    """Detecta se há registros legados onde depara.ID_FISICO aponta para fisico.row_id."""
    try:
        row = con.execute(
            """
            SELECT 1
            FROM depara d
            WHERE COALESCE(d.ID_FISICO, 0) > 0
              AND NOT EXISTS (SELECT 1 FROM fisico f WHERE f.ID = d.ID_FISICO)
              AND EXISTS (SELECT 1 FROM fisico fr WHERE fr.row_id = d.ID_FISICO)
            LIMIT 1
            """
        ).fetchone()
        return bool(row)
    except Exception:
        return False


# =========================
# BsDePara (conciliados)
# =========================

def _alias_expr_for_header(h: str) -> Tuple[str, str]:
    header = (h or "").strip()
    header_norm = re.sub(r"[\s\._-]+", "", header.upper())

    if header == "":
        return "''", ""

    if header.upper() in ("ST_CONCILIACAO", "ST. CONCILIACAO", "ST_CONCILIAÇÃO", "ST. CONCILIAÇÃO"):
        return "d.ST_CONCILIACAO", header

    # Variações comuns de campos de incorporação/incorp.
    if header_norm in ("INCCTB", "INCCONTABIL", "INCORPORACAOCTB", "INCORPORACAOCONTABIL"):
        return "COALESCE(c.INC, d.INC_CONTABIL)", header
    if header_norm in ("INCFIS", "INCORPORACAOFIS"):
        return "f.INC", header

    # mapas simples (para templates com sufixos)
    ctb_map = {
        "ID_CTB": "c.ID",
        "COD. CONTA": "c.COD_CONTA",
        "DESCR. CONTA": "c.DESC_CONTA",
        "FILIAL_CTB": "c.FILIAL",
        "DESC.FILIAL_CTB": "c.DESC_FILIAL",
        "CCUSTO_CTB": "c.CCUSTO",
        "DESCR.CCUSTO_CTB": "c.DESCR_CCUSTO",
        "LOCAL_CTB": "c.LOCAL",
        "DESCR. LOCAL_CTB": "c.DESCR_LOCAL",
        "NRBRM_CTB": "c.NRBRM",
        "INC_CTB": "c.INC",
        "DESCRICAO_CTB": "c.DESCRICAO",
        "MARCA_CTB": "c.MARCA",
        "MODELO_CTB": "c.MODELO",
        "SERIE_CTB": "c.SERIE",
        "DIMENSAO_CTB": "c.DIMENSAO",
        "CAPACIDADE_CTB": "c.CAPACIDADE",
        "TAG_CTB": "c.TAG",
        "BEM ANTERIOR_CTB": "c.BEM_ANTERIOR",
        "QTD_CTB": "c.QTD",
        "DT. AQUISIÇÃO_CTB": "c.DT_AQUISICAO",
        "VLR. AQUISIÇÃO_CTB": "c.VLR_AQUISICAO",
        "DEP. ACUMULADA_CTB": "c.DEP_ACUMULADA",
        "VLR. RESIDUAL_CTB": "c.VLR_RESIDUAL",
        "FRAG_CTB": "c.FRAG",
    }
    if header in ctb_map:
        return ctb_map[header], header

    fis_map = {
        "ID_FIS": "f.ID",
        "FILIAL_FIS": "f.FILIAL",
        "DESC.FILIAL_FIS": "f.DESC_FILIAL",
        "CCUSTO_FIS": "f.CCUSTO",
        "DESCR.CCUSTO_FIS": "f.DESCR_CCUSTO",
        "LOCAL_FIS": "f.LOCAL",
        "DESCR. LOCAL_FIS": "f.DESCR_LOCAL",
        "NRBRM_FIS": "f.NRBRM",
        "INC_FIS": "f.INC",
        "DESCRICAO_FIS": "f.DESCRICAO",
        "MARCA_FIS": "f.MARCA",
        "MODELO_FIS": "f.MODELO",
        "SERIE_FIS": "f.SERIE",
        "DIMENSAO_FIS": "f.DIMENSAO",
        "CAPACIDADE_FIS": "f.CAPACIDADE",
        "TAG_FIS": "f.TAG",
        "BEM ANTERIOR_FIS": "f.BEM_ANTERIOR",
        "CONDIC_FIS": "f.CONDIC",
        "QTD_FIS": "f.QTD",
        "FRAG_FIS": "f.FRAG",
    }
    if header in fis_map:
        return fis_map[header], header

    # fallback
    return "''", header


def export_bsdepara(
    db_path: str,
    template_xlsx: str,
    out_xlsx: str,
    *,
    sheet_name: str = "BsDePara",
    ultra_fast: bool = False,
) -> int:
    """Exporta BsDePara (conciliados) + sobras BsFisico/BsContabil mantendo o layout do template."""

    # imports pesados somente aqui
    import pandas as pd
    from openpyxl import load_workbook

    con = connect(db_path)
    try:
        # Ajustes de leitura para reduzir I/O durante o export em bases grandes.
        try:
            con.execute("PRAGMA temp_store=MEMORY;")
            con.execute("PRAGMA cache_size=-200000;")
            con.execute("CREATE INDEX IF NOT EXISTS idx_depara_idf_export ON depara(ID_FISICO);")
            con.execute("CREATE INDEX IF NOT EXISTS idx_depara_idc_export ON depara(ID_CONTABIL);")
            con.execute("CREATE INDEX IF NOT EXISTS idx_fis_id_export ON fisico(ID);")
            con.execute("CREATE INDEX IF NOT EXISTS idx_ctb_id_export ON contabil(ID);")
        except Exception:
            pass

        wb = load_workbook(template_xlsx)

        template_headers = _read_headers(wb, sheet_name)
        if not template_headers:
            raise ValueError(f"Não encontrei cabeçalhos na aba '{sheet_name}' do template.")

        select_parts: List[str] = []
        for h in template_headers:
            expr, alias = _alias_expr_for_header(h)
            if alias == "":
                select_parts.append(f"{expr} AS \"\"")
            else:
                select_parts.append(f"{expr} AS \"{alias}\"")

        select_clause = ",\n            ".join(select_parts)
        has_fis_row_id = _table_has_column(con, "fisico", "row_id")
        use_rowid_fallback = has_fis_row_id and _needs_fis_rowid_fallback(con)
        fis_join_cond = "(f.ID = d.ID_FISICO OR f.row_id = d.ID_FISICO)" if use_rowid_fallback else "f.ID = d.ID_FISICO"

        join_q = (
            "SELECT\n            " + select_clause + "\n"
            "FROM depara d\n"
            "LEFT JOIN contabil c ON c.ID = d.ID_CONTABIL\n"
            f"LEFT JOIN fisico   f ON {fis_join_cond}\n"
            "ORDER BY d.rowid"
        )

        out = pd.read_sql_query(join_q, con)

        for col in out.columns:
            if str(col).strip().upper().startswith("DT. AQUISIÇÃO"):
                def _fmt_date(v):
                    if pd.isna(v) or v == "":
                        return ""
                    try:
                        return pd.to_datetime(v).strftime("%d/%m/%Y")
                    except Exception:
                        return str(v)
                out[col] = out[col].apply(_fmt_date)

        fis_pend = _build_pending_df(con, "FIS")
        ctb_pend = _build_pending_df(con, "CTB")

        ws_de = _ensure_sheet(wb, sheet_name, template_headers)
        ws_fis = _ensure_sheet(wb, "BsFisico", FISICO_HEADERS)
        ws_ctb = _ensure_sheet(wb, "BsContabil", CONTABIL_HEADERS)

        _write_df(ws_de, out, template_headers)
        _write_df(ws_fis, fis_pend, FISICO_HEADERS)
        _write_df(ws_ctb, ctb_pend, CONTABIL_HEADERS)

        _save_workbook_fast(wb, out_xlsx, compresslevel=1, ultra_fast=ultra_fast)
        return int(len(out))
    finally:
        con.close()


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Exportador V2 (lazy imports)")
    p.add_argument("--db", default="conciliador.db")
    p.add_argument("--template", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--ultra-fast", action="store_true", help="Prioriza velocidade máxima de export (arquivo maior).")
    args = p.parse_args()

    n = export_bsdepara(args.db, args.template, args.out, ultra_fast=args.ultra_fast)
    print(f"Exportados {n} pares para BsDePara (e sobras para BsFisico/BsContabil).")
