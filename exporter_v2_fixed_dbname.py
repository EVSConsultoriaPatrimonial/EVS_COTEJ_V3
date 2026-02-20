# exporter_v2.py
from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


# Cabeçalhos do template (BsDePara.xlsx) — mantidos como no seu modelo
TEMPLATE_HEADERS = [
    "ID_CTB","FILIAL_CTB","DESC.FILIAL_CTB","CCUSTO_CTB","DESCR.CCUSTO_CTB","LOCAL_CTB","DESCR. LOCAL_CTB",
    "NRBRM_CTB","INC._CTB","DESCRICAO_CTB","MARCA_CTB","MODELO_CTB","SERIE_CTB","DIMENSAO_CTB","CAPACIDADE_CTB",
    "TAG_CTB","BEM ANTERIOR_CTB","QTD_CTB","DT. AQUISIÇÃO_CTB","VLR. AQUISIÇÃO_CTB","DEP. ACUMULADA_CTB","VLR. RESIDUAL_CTB",
    "FRAG_CTB","ST_CONCILIACAO",
    "ID_FIS","FILIAL_FIS","DESC.FILIAL_FIS","CCUSTO_FIS","DESCR.CCUSTO_FIS","LOCAL_FIS","DESCR. LOCAL_FIS",
    "NRBRM_FIS","INC._FIS","DESCRICAO_FIS","MARCA_FIS","MODELO_FIS","SERIE_FIS","DIMENSAO_FIS","CAPACIDADE_FIS",
    "TAG_FIS","BEM ANTERIOR_FIS","CONDIC_FIS","QTD_FIS","FRAG_FIS"
]


def export_bsdepara(db_path: str, template_xlsx: str, out_xlsx: str, *, sheet_name: str = "BsDePara") -> int:
    """
    Exporta o resultado do SQLite para uma planilha BsDePara seguindo o layout do seu modelo.
    - Mantém o template (formatos/cabeçalho) e escreve a partir da linha 2.
    - FRAG_CTB e FRAG_FIS recebem o PAR_ID (sequencial) para referência no arquivo final.
    """
    con = sqlite3.connect(db_path)
    try:
        q = """
        SELECT d.PAR_ID, d.ST_CONCILIACAO, d.ID_FISICO, d.ID_CONTABIL
        FROM depara d
        ORDER BY d.PAR_ID;
        """
        de = pd.read_sql_query(q, con)
        if de.empty:
            # ainda assim gera arquivo baseado no template
            wb = load_workbook(template_xlsx)
            wb.save(out_xlsx)
            return 0

        fis = pd.read_sql_query("SELECT * FROM fisico;", con)
        ctb = pd.read_sql_query("SELECT * FROM contabil;", con)

        fis = fis.rename(columns={
            "ID":"ID_FIS", "FILIAL":"FILIAL_FIS", "DESC_FILIAL":"DESC.FILIAL_FIS",
            "CCUSTO":"CCUSTO_FIS", "DESCR_CCUSTO":"DESCR.CCUSTO_FIS",
            "LOCAL":"LOCAL_FIS", "DESCR_LOCAL":"DESCR. LOCAL_FIS",
            "NRBRM":"NRBRM_FIS", "INC":"INC._FIS", "DESCRICAO":"DESCRICAO_FIS",
            "MARCA":"MARCA_FIS", "MODELO":"MODELO_FIS", "SERIE":"SERIE_FIS",
            "DIMENSAO":"DIMENSAO_FIS", "CAPACIDADE":"CAPACIDADE_FIS", "TAG":"TAG_FIS",
            "BEM_ANTERIOR":"BEM ANTERIOR_FIS", "CONDIC":"CONDIC_FIS", "QTD":"QTD_FIS",
        })

        ctb = ctb.rename(columns={
            "ID":"ID_CTB", "FILIAL":"FILIAL_CTB", "DESC_FILIAL":"DESC.FILIAL_CTB",
            "CCUSTO":"CCUSTO_CTB", "DESCR_CCUSTO":"DESCR.CCUSTO_CTB",
            "LOCAL":"LOCAL_CTB", "DESCR_LOCAL":"DESCR. LOCAL_CTB",
            "NRBRM":"NRBRM_CTB", "INC":"INC._CTB", "DESCRICAO":"DESCRICAO_CTB",
            "MARCA":"MARCA_CTB", "MODELO":"MODELO_CTB", "SERIE":"SERIE_CTB",
            "DIMENSAO":"DIMENSAO_CTB", "CAPACIDADE":"CAPACIDADE_CTB", "TAG":"TAG_CTB",
            "BEM_ANTERIOR":"BEM ANTERIOR_CTB", "QTD":"QTD_CTB",
            "DT_AQUISICAO":"DT. AQUISIÇÃO_CTB", "VLR_AQUISICAO":"VLR. AQUISIÇÃO_CTB",
            "DEP_ACUMULADA":"DEP. ACUMULADA_CTB", "VLR_RESIDUAL":"VLR. RESIDUAL_CTB",
        })

        out = de.merge(ctb, left_on="ID_CONTABIL", right_on="ID_CTB", how="left")
        out = out.merge(fis, left_on="ID_FISICO", right_on="ID_FIS", how="left")

        out["FRAG_CTB"] = out["PAR_ID"]
        out["FRAG_FIS"] = out["PAR_ID"]
        out["ST_CONCILIACAO"] = out["ST_CONCILIACAO"].fillna("")

        # Monta exatamente as colunas do template
        for col in TEMPLATE_HEADERS:
            if col not in out.columns:
                out[col] = ""

        out = out[TEMPLATE_HEADERS]

        # --- escreve no template preservando formatos
        wb = load_workbook(template_xlsx)

        # Seleção flexível da aba (evita erro quando o usuário escolhe o arquivo errado ou o nome varia)
        _sheet = None
        # match exato
        if sheet_name in wb.sheetnames:
            _sheet = sheet_name
        else:
            # match case-insensitive e ignorando espaços
            wanted = sheet_name.strip().lower()
            for s in wb.sheetnames:
                if s.strip().lower() == wanted:
                    _sheet = s
                    break
        if _sheet is None:
            # Se existir apenas 1 aba, usa ela. Caso contrário usa a primeira e informa as opções na mensagem.
            if len(wb.sheetnames) == 1:
                _sheet = wb.sheetnames[0]
            else:
                _sheet = wb.sheetnames[0]
        ws = wb[_sheet]

        # Limpa linhas antigas (a partir da 2) sem mexer no cabeçalho
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Escreve
        start_row = 2
        for i, row in enumerate(out.itertuples(index=False), start=start_row):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)

        wb.save(out_xlsx)
        return int(len(out))
    finally:
        con.close()


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Exportador V2 -> BsDePara.xlsx (layout do modelo)")
    p.add_argument("--db", default="conciliador.db")
    p.add_argument("--template", required=True, help="Template BsDePara.xlsx (modelo)")
    p.add_argument("--out", required=True, help="Arquivo BsDePara de saída")
    args = p.parse_args()

    n = export_bsdepara(args.db, args.template, args.out)
    print(f"OK: exportado {n} pares para {args.out}")
